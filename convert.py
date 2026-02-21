import pandas as pd
import sys
import os

def converter_csv_para_excel(arquivo_csv, arquivo_saida=None):
    # Tenta diferentes encodings comuns em exports do Chatwoot
    encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252']
    
    df = None
    for enc in encodings:
        try:
            df = pd.read_csv(arquivo_csv, encoding=enc)
            print(f"Arquivo lido com encoding: {enc}")
            break
        except (UnicodeDecodeError, Exception):
            continue
    
    if df is None:
        print("Erro: não foi possível ler o arquivo CSV.")
        return

    # Corrige caracteres mal codificados (ex: latin-1 interpretado como utf-8)
    for col in df.select_dtypes(include='object').columns:
        try:
            df[col] = df[col].apply(lambda x: 
                x.encode('latin-1').decode('utf-8') if isinstance(x, str) else x
            )
        except (UnicodeDecodeError, UnicodeEncodeError):
            pass  # Coluna já está correta

    if arquivo_saida is None:
        base = os.path.splitext(arquivo_csv)[0]
        arquivo_saida = base + '.xlsx'

    # Salva como Excel com formatação básica
    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Contatos')
        
        ws = writer.sheets['Contatos']
        
        # Formata cabeçalho
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill('solid', start_color='2B5ED6')
        header_font = Font(bold=True, color='FFFFFF', name='Arial')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Ajusta largura das colunas
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        
        # Fonte padrão para dados
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name='Arial', size=10)

    print(f"\n✅ Arquivo salvo com sucesso: {arquivo_saida}")
    print(f"   Total de contatos: {len(df)}")
    print(f"   Colunas: {', '.join(df.columns.tolist())}")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python converter_chatwoot.py arquivo.csv [saida.xlsx]")
        print("\nEste script:")
        print("  1. Corrige caracteres especiais (acentos, ç, etc.)")
        print("  2. Converte o CSV para planilha Excel formatada")
    else:
        saida = sys.argv[2] if len(sys.argv) > 2 else None
        converter_csv_para_excel(sys.argv[1], saida)